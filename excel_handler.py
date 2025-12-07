import shutil
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import config
import time


class ExcelHandler:

    STATUS_MAP = {
        "Director_confirm_form": "ФІНДИРЕКТОР",
        "Financial_namager_confirm_form": "ДИРЕКТОР",
        "Empty_form": "ДИРЕКТОР",
    }

    def is_file_locked(self, file_path: str) -> bool:
        try:
            with open(file_path, "a"):
                return False
        except (PermissionError, OSError, IOError):
            return True

    def read_application(self, file_path):
        wb = None
        try:
            if self.is_file_locked(file_path):
                return None

            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            if "Бланк" not in wb.sheetnames or "Налаштування" not in wb.sheetnames:
                return None

            blank = wb["Бланк"]
            settings = wb["Налаштування"]
            
            status = settings["B8"].value
            approver = self.STATUS_MAP.get(status)

            if approver is None:
                return None

            raw_date = blank["B1"].value
            date_str = (raw_date.strftime("%d.%m.%Y") if hasattr(raw_date, "strftime")
                        else str(raw_date or "—").strip())

            suma = blank["B10"].value or 0
            try:
                suma_str = f"{float(suma):,.2f}".replace(",", " ") + " грн"
            except:
                suma_str = f"{suma} грн"

            payment_raw = blank["C3"].value or ""

            data = {
                "file_path": str(Path(file_path).resolve()),
                "file_name": Path(file_path).name,
                "дата": date_str,
                "заявник": blank["E1"].value or "—",
                "відділ": blank["H1"].value or "—",
                "сума": suma_str,
                "постачальник": blank["G4"].value or "—",
                "призначення": blank["C12"].value or "—",
                "вид_розрахунку": payment_raw,
                "intended_approver": approver,
                "статус": status,
            }
            
            return data

        except Exception as e:
            print(f"Помилка читання {Path(file_path).name}: {e}")
            return None
        
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass

    def move_file(self, file_path, approved=True):
        src = Path(file_path)
        
        if not src.exists():
            return True

    
        if not approved:
            dest_folder = config.get_path("rejected_folder")
            if not dest_folder:
                return False
            dest_path = Path(dest_folder) / src.name
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            if dest_path.exists():
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest_path = dest_path.parent / f"{src.stem}_{ts}{src.suffix}"
            return self._safe_move(src, dest_path)


        try:
            wb = load_workbook(str(src), data_only=True, read_only=True)
            status = wb["Налаштування"]["B8"].value
            payment_raw = str(wb["Бланк"]["C3"].value or "").strip().upper()
            wb.close()
        except Exception as e:
            print(f"❌ Помилка читання файлу при переміщенні: {e}")
            return False


        current_folder = src.parent.resolve()
    
        findirector_folder = config.get_path("findirector_folder")
        director_folder = config.get_path("director_folder")
        
        if not findirector_folder or not director_folder:
            print("❌ Папки не налаштовані")
            return False
            
        findirector_path = Path(findirector_folder).resolve()
        director_path = Path(director_folder).resolve()
        
        print(f"📍 Поточна папка: {current_folder}")
        print(f"📋 Статус у файлі: {status}")
        print(f"💳 Вид розрахунку: {payment_raw}")
        
    
        if current_folder == findirector_path:
            dest_folder = director_folder
            print(f"➡️ Маршрут: Фіндиректор → Директор")
            
        elif current_folder == director_path:
    
            if any(kw in payment_raw for kw in ["БЕЗГОТІВКА", "КАРТА", "КАРТКА"]):
                dest_folder = config.get_path("accountant_folder")
                print(f"➡️ Маршрут: Директор → Бухгалтер (безготівка)")
            else:
                dest_folder = config.get_path("cashier_folder")
                print(f"➡️ Маршрут: Директор → Касир (готівка)")
        else:

            print(f"⚠️ Файл у невідомій папці, використовуємо логіку за статусом")
            if status == "Director_confirm_form":
                dest_folder = director_folder
            elif status in ("Financial_namager_confirm_form", "Empty_form"):
                if any(kw in payment_raw for kw in ["БЕЗГОТІВКА", "КАРТА", "КАРТКА"]):
                    dest_folder = config.get_path("accountant_folder")
                else:
                    dest_folder = config.get_path("cashier_folder")
            else:
                print(f"❌ Невідомий статус: {status}")
                return False

        if not dest_folder:
            print("❌ Цільова папка не налаштована")
            return False

        dest_path = Path(dest_folder) / src.name
        
        if dest_path.resolve() == src.resolve():
            print(f"⚠️ Файл вже у цільовій папці: {src.name}")
            return True
        
        print(f"🎯 Цільова папка: {dest_path.parent}")
        
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        
    
        if dest_path.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest_path = dest_path.parent / f"{src.stem}_{ts}{src.suffix}"
            print(f"⚠️ Файл існує, додано timestamp: {dest_path.name}")

        return self._safe_move(src, dest_path)

    def _safe_move(self, src: Path, dst: Path) -> bool:
        try:
            
            shutil.copy2(str(src), str(dst))
            print(f"✅ Скопійовано → {dst.parent.name}/{dst.name}")

    
            for attempt in range(10):
                try:
                    src.unlink()
                    print(f"🗑️ Оригінал видалено")
                    return True
                except PermissionError:
                    time.sleep(1)
                except FileNotFoundError:
    
                    return True

            print("⚠️ Оригінал залишено (відкритий у Excel), але копія створена")
            return True

        except Exception as e:
            print(f"❌ Критична помилка переміщення: {e}")
            return False

    def validate_file(self, file_path):
        wb = None
        try:
            if not file_path.lower().endswith(('.xlsm', '.xlsx')):
                return False, "Непідтримуваний формат"

            if not Path(file_path).exists():
                return False, "Файл не знайдено"

            wb = load_workbook(file_path, data_only=True, read_only=True)
            has_sheets = all(sh in wb.sheetnames for sh in ["Бланк", "Налаштування"])
            
            return (True, "OK") if has_sheets else (False, "Немає потрібних аркушів")
            
        except Exception as e:
            return False, f"Помилка: {e}"
        
        finally:
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass

